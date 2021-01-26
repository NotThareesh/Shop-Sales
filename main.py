from openpyxl import load_workbook
import sys
import time

wb = load_workbook("customers.xlsx")
sheet = wb.active

current_choice = ""
available_names = []
available_prices = []
names = []
balances = []
cart = []
cart_total = 0
count = 0

# Reading all the cells in the mentioned file
for col in sheet.iter_cols(max_col=1, min_row=2):
    for cell in col:
        names.append(cell.value)

for col in sheet.iter_cols(min_col=3, min_row=2):
    for cell in col:
        balances.append(cell.value)


def check_balance():
    index_cell_1 = names.index(name_prompt) + 2
    if balances[names.index(name_prompt)] >= cart_total:
        balance = int(balances[names.index(name_prompt)] - cart_total)
        sheet[f"C{index_cell_1}"] = balance
        print(f"${cart_total} was credited from your account")
        print(f"Remaining Balance is $ {balance:,}")
        wb.save("customers.xlsx")
    else:
        print("Insufficient Balance in your account")


def get_items():
    print("Please add options from the list below: ")
    for number, part in enumerate(available_names):
        print("{0}: {1}".format(number + 1, part))

# Reading all the contents of the .txt file in one-go
with open("items.txt", "r") as items_file:
    items_file_read = items_file.readlines()

# Using ':' as a separator, and using the index of it as a reference to separate the name and the price
for word in items_file_read:
    if ":" in word:
        index = word.index(":")

        # No Price Mentioned for the Product
        if word[index + 2:len(word)] == "":
            print("No price mentioned for {}".format(word[0:index]))
            sys.exit()

        # No Name Mentioned for the Product
        elif word[0:index] == "":
            print("No name mentioned for product after {}".format(available_names[-1]))
            sys.exit()

        else:
            available_names.append(word[0:index])  # Appending Names
            if word[index + 1] in "123456789":
                available_prices.append(int(word[index + 1:len(word)]))  # Appending the prices
            else:
                available_prices.append(int(word[index + 2:len(word)]))

    else:
        print("No ':' mentioned for the product after {}".format(available_names[-1]))
        sys.exit()

# Making a dictionary comprised of the contents of 'available_names' and 'available_prices'
zip_obj = zip(available_names, available_prices)
available_names_prices = dict(zip_obj)

# Making a list of all the valid choices by iterating through the 'available_names' list
valid_choices = [str(i) for i in range(1, len(available_names) + 1)]

if __name__ == "__main__":

    # Printing all the items initially
    get_items()

    while current_choice != "0":
        count += 1

        if current_choice in available_names:
            cart.append(current_choice)
            print(f"{current_choice} Added")

        elif current_choice in valid_choices:
            chosen_part = available_names[int(current_choice) - 1]
            cart.append(chosen_part)
            print(f"{chosen_part} Added")

        elif current_choice.lower() == "cart":
            new_list = []

            if len(cart) >= 1:
                print("These are the items in your cart: ")

                for items in cart:
                    cart_total += available_names_prices[items]

                    if items not in new_list:
                        new_list.append(items)

                        if cart.count(items) > 1:
                            print(f"{items} * {cart.count(items)}")

                        else:
                            print(items)

                print(f"\nThe total cost is ${cart_total}")

            else:
                print("The cart is empty. Please add some items.")

        elif current_choice.lower() in ("list", "options", "items"):
            get_items()

        elif current_choice.lower() in ("buy", "checkout"):
            index_sheet = sheet.max_row + 1
            if len(cart) >= 1:

                for items in cart:
                    cart_total += available_names_prices[items]

                name_prompt = input("Please enter your name: ")
                index_cell = names.index(name_prompt) + 2
                password_prompt = input("Please enter your password: ")

                if str(sheet[f"B{index_cell}"].value) == password_prompt:
                    check_balance()

                elif name_prompt not in names:
                    print("Account Not Found")
                    prompt = input("Would you like to create a new account? (Yes/No): ")

                    if prompt.lower() in ("yes", "y"):
                        # Creating a new account in the Excel Sheet
                        name_prompt = input("Please enter your name: ")
                        balance_prompt = int(input("Please enter the balance: "))
                        password_prompt = input("Please enter the password: ")
                        sheet[f"A{index_sheet}"] = name_prompt
                        sheet[f"C{index_sheet}"] = balance_prompt
                        names.append(name_prompt)
                        balances.append(balance_prompt)
                        print("Account Created.\nName is {0}, Balance = {1}\n".format(names[-1], balances[-1]))

                        check_balance()

                    else:
                        print("You cannot make a purchase without an account")
                        time.sleep(3)
                        sys.exit()

                else:
                    print("Incorrect password")
                    sys.exit()

                time.sleep(5)
                sys.exit()

            else:
                print("There's nothing in the cart.")

        elif current_choice not in available_names and valid_choices:
            if count > 1:
                print("Item Not Found. Please be more specific")

        current_choice = input("\nEnter your choice: ")
